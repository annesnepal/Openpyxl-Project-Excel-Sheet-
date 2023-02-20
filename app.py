import openpyxl as xl
from openpyxl.chart import BarChart, Reference
def work_book(filename):
    wb = xl.load_workbook('transactions.xlsx')
    sheet = wb['Sheet1']
    sheet['d1'] = "updated_price"
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.8
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
            min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'a15')

    wb.save(filename)

work_book('transactions.xlsx')

